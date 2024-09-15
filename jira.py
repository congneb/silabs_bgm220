from openpyxl import load_workbook
import os
import sys
import re
import json

# Github API got issue
def github_api_issue(repo_name, ISSUE_NUMBER):
  REPO_URL = "SiliconLabs/" + repo_name
  AN_ACCESS_KEY = os.environ.get('AN_ACCESS_KEY_PSW')
  
  cmd_got_issue = "curl -L -H \"Accept: application/vnd.github+json\" -H \"Authorization: Bearer " + AN_ACCESS_KEY + "\" -H \"X-GitHub-Api-Version: 2022-11-28\" https://api.github.com/repos/" + REPO_URL + "/issues/" + ISSUE_NUMBER + " > got_gh_issue.json"
  os.system(cmd_got_issue)	

# If a new issue be 
def add_new_issue_to_excel(workbook):

	sheet = workbook.active

	# Data to be added
	new_row = ['Data1', 'Data2', 'Data3']

	# Append the new row
	sheet.append(new_row)

	# Save the workbook
	workbook.save('your_file.xlsx')

# Delete 
def remove_issue_from_excel(workbook, row_id):
	# Load the workbook and select the active worksheet
	workbook = load_workbook('your_file.xlsx')
	sheet = workbook.active

	# Delete the row
	sheet.delete_rows(row_id)

	# Save the workbook
	workbook.save('your_file.xlsx')

# Scan excel file
# Nếu issue đã close => Remove trong file excel
def scan_excel(workbook):
	repo_name = sys.argv[1]
	sheet_name = repo_name
	sheet = workbook[repo_name]

	# Get the last row number
	# Cong: Có 1 issue, nếu có dấu Space hay đóng khung Cell thì nó vẫn tính
	last_row = sheet.max_row
	print("Last Row Number:", last_row)


	# Scan all row in each Sheet
	for row_id in range(1, last_row):
		issue_number = sheet.cell(row = row_id, column = 2).value
		print("Scanning for issue number: ", issue_number)

		github_api_issue(repo_name, issue_number)
		# Check if this issue be Closed then remove this issue from excel file
		json_path = "got_gh_issue.json"
		with open(json_path, "r") as json_file:
			data = json.load(json_file)

		issue_status = data[id]['status']
		if issue_status == "closed":
			remove_issue_from_excel(workbook, row_id)


def scan_excel_v1(workbook):
	repo_name = sys.argv[1]
	sheet_name = repo_name
	sheet = workbook[repo_name]

	# Get the last row number
	# last_row = sheet.max_row
	# print("Last Row Number:", last_row)

	# way2
	for i in range(1, 50):
		no_data = sheet.cell(row = i, column = 1).value
		issue_number = sheet.cell(row = i, column = 2).value
		if no_data == None and issue_number == None:
			last_row = i - 1
			print("Last Row Number:", last_row)
			return None

####################################

# Load the existing Excel file
workbook = load_workbook('your_file.xlsx')

# add_new_issue(workbook)

# Step 1. Scan all issue
scan_excel_v1(workbook)

# Step 2. Check new Issue or not
